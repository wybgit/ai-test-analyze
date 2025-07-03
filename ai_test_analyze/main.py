import argparse
import json
import os
import requests
import csv
from datetime import datetime
from pathlib import Path
import fnmatch
from tqdm import tqdm
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import time
import sys
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
import shutil
import re

# --- 全局常量 ---
CONFIG_DIR = Path.home() / ".config" / "ai-test-analyze"
TEMPLATES_DIR_IN_USER = CONFIG_DIR / "templates"
CONFIG_FILE = CONFIG_DIR / "config.json"
PROMPT_TEMPLATE_FILE = TEMPLATES_DIR_IN_USER / "prompt.template"
SUCCESS_PATTERN_FILE = TEMPLATES_DIR_IN_USER / "success_pattern.template"
FAILED_PATTERN_FILE = TEMPLATES_DIR_IN_USER / "failed_pattern.template"
EXCEPTION_PATTERN_FILE = TEMPLATES_DIR_IN_USER / "exception_pattern.template"
LOG_READ_BUFFER_SIZE = 8192 # Read last 8KB of log files
SAVE_INTERVAL = 100 # Save the Excel file every 100 records

# --- 辅助与配置函数 ---
def get_templates_dir_in_pkg():
    return Path(__file__).parent / "templates"

def ensure_config_files_exist():
    CONFIG_DIR.mkdir(parents=True, exist_ok=True)
    TEMPLATES_DIR_IN_USER.mkdir(exist_ok=True)
    
    pkg_templates_dir = get_templates_dir_in_pkg()

    if not CONFIG_FILE.exists():
        default_config = {
            "api_token": "your-api-key-here", "api_url": "https://api.siliconflow.cn/v1/chat/completions",
            "model": "Qwen/Qwen3-235B-A22B", "max_tokens": 8192, "temperature": 0.6, "top_p": 0.7,
            "fast_success_regex": ["All tests passed successfully", "^\\s*OK\\s*$", "BUILD SUCCESSFUL"],
            "fast_failure_regex": ["(?i)ERROR", "(?i)FATAL", "Traceback \\(most recent call last\\):", "BUILD FAILED"],
            "directory_whitelist": ["*"], "logfile_whitelist": ["*.log", "*.txt"],
            "directory_blacklist": [".git", "__pycache__"], "logfile_blacklist": []
        }
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(default_config, f, indent=4, ensure_ascii=False)
        print(f"默认配置文件已创建于: {CONFIG_FILE}")

    template_files = ["prompt.template", "success_pattern.template", "failed_pattern.template", "exception_pattern.template"]
    for template_file in template_files:
        user_path = TEMPLATES_DIR_IN_USER / template_file
        pkg_path = pkg_templates_dir / template_file
        if not user_path.exists() and pkg_path.exists():
            shutil.copy(pkg_path, user_path)
            print(f"默认模板 '{template_file}' 已创建于: {user_path}")

def load_templates():
    with open(PROMPT_TEMPLATE_FILE, 'r', encoding='utf-8') as f: prompt_template = f.read()
    with open(SUCCESS_PATTERN_FILE, 'r', encoding='utf-8') as f: success_patterns = f.read()
    with open(FAILED_PATTERN_FILE, 'r', encoding='utf-8') as f: failed_patterns = f.read()
    with open(EXCEPTION_PATTERN_FILE, 'r', encoding='utf-8') as f: exception_patterns = f.read()
    return prompt_template, success_patterns, failed_patterns, exception_patterns

# ... (The rest of the main.py file remains largely the same)
STATIC_HEADERS_PRE = ["Root Dir"]
STATIC_HEADERS_POST = ["Log File Name", "Absolute Path", "Analysis Result", "Analysis Details", "Extracted Log Content"]
DEBUG_HEADERS = ["Final Prompt to LLM", "LLM Reasoning & Response"]
STATUS_PENDING = "Pending"; STATUS_SUCCESS = "成功"; STATUS_FAILURE = "失败"
CSV_ENCODING = 'utf-8-sig'

FILL_SUCCESS = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_FAILURE = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_ERROR = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FONT_BOLD = Font(bold=True)
ALIGNMENT_WRAP = Alignment(wrap_text=True, vertical='top')

def find_log_files(target_path, config):
    log_files = []
    base_path = Path(target_path).resolve()
    dir_w = config.get("directory_whitelist", ["*"])
    file_w = config.get("logfile_whitelist", ["*.log"])
    dir_b = config.get("directory_blacklist", [])
    file_b = config.get("logfile_blacklist", [])
    
    all_paths = list(os.walk(base_path, topdown=True))
    
    for root, dirs, files in tqdm(all_paths, desc="Scanning Directories"):
        dirs[:] = [d for d in dirs if any(fnmatch.fnmatch(d, p) for p in dir_w) and not any(fnmatch.fnmatch(d, p) for p in dir_b)]
        for file in files:
            if any(fnmatch.fnmatch(file, p) for p in file_w) and not any(fnmatch.fnmatch(file, p) for p in file_b):
                log_files.append(str(Path(root) / file))
    return log_files

def get_dynamic_headers(log_files, base_path_str, debug):
    max_depth = 0
    base_path = Path(base_path_str).resolve()
    for log_file in log_files:
        depth = len(Path(log_file).relative_to(base_path).parts) - 1
        if depth > max_depth: max_depth = depth
    
    dynamic_dir_headers = [f"Sub Dir {i+1}" for i in range(max_depth)]
    headers = STATIC_HEADERS_PRE + dynamic_dir_headers + STATIC_HEADERS_POST
    if debug:
        headers.extend(DEBUG_HEADERS)
    return headers, max_depth

def get_dir_parts(log_path_str, base_path_str, max_depth):
    base_path = Path(base_path_str).resolve()
    log_path = Path(log_path_str).resolve()
    relative_path = log_path.relative_to(base_path)
    dir_parts = list(relative_path.parent.parts)
    dir_parts.extend([''] * (max_depth - len(dir_parts)))
    return [base_path.name] + dir_parts, relative_path.name

def initialize_report(report_path, headers, log_files, max_depth, base_path):
    ext = Path(report_path).suffix
    if ext == '.csv':
        _initialize_csv(report_path, headers, log_files, max_depth, base_path)
    elif ext == '.xlsx':
        _initialize_xlsx(report_path, headers, log_files, max_depth, base_path)

def get_tasks_from_report(report_path):
    ext = Path(report_path).suffix
    if ext == '.csv':
        return _get_tasks_from_csv(report_path)
    elif ext == '.xlsx':
        return _get_tasks_from_xlsx(report_path)
    return [], None, []

def update_report_row(report_path, row_index, data, headers, lock, wb=None):
    with lock:
        ext = Path(report_path).suffix
        if ext == '.csv':
            _update_csv_row(report_path, row_index, data, headers)
        elif ext == '.xlsx' and wb:
            _update_xlsx_row(wb.active, row_index, data, headers)

def finalize_report(report_path, headers, wb=None):
    if Path(report_path).suffix == '.xlsx' and wb:
        _finalize_xlsx(wb, headers, report_path)

def _initialize_csv(path, headers, logs, max_depth, base):
    with open(path, 'w', newline='', encoding=CSV_ENCODING) as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for log in logs:
            dir_parts, file_name = get_dir_parts(log, base, max_depth)
            row = dir_parts + [file_name, log, STATUS_PENDING, "", ""]
            if "Final Prompt to LLM" in headers:
                row.extend(["", ""])
            writer.writerow(row)
    print(f"任务列表已生成: {path}。共 {len(logs)} 个文件。")

def _get_tasks_from_csv(path):
    with open(path, 'r', encoding=CSV_ENCODING) as f:
        reader = csv.reader(f)
        headers = next(reader)
        path_idx = headers.index("Absolute Path")
        result_idx = headers.index("Analysis Result")
        tasks = []
        for i, row in enumerate(reader):
            if row[result_idx] == STATUS_PENDING:
                tasks.append((i + 2, row[path_idx]))
    return tasks, None, headers

def _update_csv_row(path, index, data, headers):
    rows = list(csv.reader(open(path, 'r', encoding=CSV_ENCODING)))
    res, reason, log_content, req, resp = data
    res_idx = headers.index("Analysis Result")
    det_idx = headers.index("Analysis Details")
    log_idx = headers.index("Extracted Log Content")
    rows[index-1][res_idx], rows[index-1][det_idx], rows[index-1][log_idx] = res, reason, log_content
    if "Final Prompt to LLM" in headers:
        req_idx = headers.index("Final Prompt to LLM")
        resp_idx = headers.index("LLM Reasoning & Response")
        rows[index-1][req_idx] = req
        rows[index-1][resp_idx] = resp
    writer = csv.writer(open(path, 'w', newline='', encoding=CSV_ENCODING))
    writer.writerows(rows)

def _initialize_xlsx(path, headers, logs, max_depth, base):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "分析报告"
    ws.append(headers)
    for cell in ws[1]: cell.font = FONT_BOLD
    for log in logs:
        dir_parts, file_name = get_dir_parts(log, base, max_depth)
        row = dir_parts + [file_name, log, STATUS_PENDING, "", ""]
        if "Final Prompt to LLM" in headers:
            row.extend(["", ""])
        ws.append(row)
    ws.freeze_panes = 'A2'
    wb.save(path)

def _get_tasks_from_xlsx(path):
    wb = openpyxl.load_workbook(path); ws = wb.active
    headers = [cell.value for cell in ws[1]]
    tasks = []
    path_idx = headers.index("Absolute Path") + 1
    result_idx = headers.index("Analysis Result") + 1
    for i, row in enumerate(ws.iter_rows(min_row=2)):
        if row[result_idx-1].value == STATUS_PENDING:
            tasks.append((i + 2, row[path_idx-1].value))
    return tasks, wb, headers

def _update_xlsx_row(ws, index, data, headers):
    res, reason, log_content, req, resp = data
    res_idx = headers.index("Analysis Result") + 1
    det_idx = headers.index("Analysis Details") + 1
    log_idx = headers.index("Extracted Log Content") + 1
    ws.cell(row=index, column=res_idx, value=res)
    ws.cell(row=index, column=det_idx, value=reason)
    ws.cell(row=index, column=log_idx, value=log_content)
    cell_to_format = ws.cell(row=index, column=res_idx)
    if res == STATUS_SUCCESS: cell_to_format.fill = FILL_SUCCESS
    elif res == STATUS_FAILURE: cell_to_format.fill = FILL_FAILURE
    else: cell_to_format.fill = FILL_ERROR
    if "Final Prompt to LLM" in headers:
        req_idx = headers.index("Final Prompt to LLM") + 1
        resp_idx = headers.index("LLM Reasoning & Response") + 1
        ws.cell(row=index, column=req_idx, value=req)
        ws.cell(row=index, column=resp_idx, value=resp)

def _finalize_xlsx(wb, headers, path):
    ws = wb.active
    abs_path_col_idx = headers.index("Absolute Path") + 1
    ws.column_dimensions[openpyxl.utils.get_column_letter(abs_path_col_idx)].hidden = True
    for col in ws.columns:
        max_len = 0; col_letter = col[0].column_letter
        for cell in col:
            cell.alignment = ALIGNMENT_WRAP
            if len(str(cell.value)) > max_len: max_len = len(str(cell.value))
        ws.column_dimensions[col_letter].width = min((max_len + 2) * 1.2, 60)
    wb.save(path)

def analyze_log(log_path, config, templates):
    prompt_template, success_patterns, failed_patterns, exception_patterns = templates
    try:
        with open(log_path, 'rb') as f:
            f.seek(-LOG_READ_BUFFER_SIZE, os.SEEK_END)
            log_content = f.read().decode('utf-8', errors='ignore')
    except (IOError, OSError):
        with open(log_path, 'r', encoding='utf-8', errors='ignore') as f:
            log_content = f.read()
    except Exception as e:
        return "文件读取错误", f"无法读取: {e}", "", "", ""

    # --- Fast Path Regex Check ---
    success_regex_list = config.get("fast_success_regex", [])
    if success_regex_list:
        for pattern in success_regex_list:
            if re.search(pattern, log_content):
                return STATUS_SUCCESS, f"Matched fast success regex: '{pattern}'", log_content, "N/A (Fast Path)", "N/A (Fast Path)"
    
    failure_regex_list = config.get("fast_failure_regex", [])
    if failure_regex_list:
        for pattern in failure_regex_list:
            if re.search(pattern, log_content):
                return STATUS_FAILURE, f"Matched fast failure regex: '{pattern}'", log_content, "N/A (Fast Path)", "N/A (Fast Path)"

    # --- LLM Analysis Path ---
    prompt = prompt_template.format(
        log_content=log_content,
        success_patterns=success_patterns,
        failed_patterns=failed_patterns,
        exception_patterns=exception_patterns
    )
    headers = {"Authorization": f"Bearer {config['api_token']}", "Content-Type": "application/json"}
    payload = {"model": config.get("model"), "messages": [{"role": "user", "content": prompt}], "max_tokens": config.get("max_tokens"), "stream": True}
    
    buffer, reasoning_content, final_content = "", "", ""
    try:
        with requests.post(config['api_url'], headers=headers, json=payload, timeout=120, stream=True) as r:
            r.raise_for_status()
            for chunk in r.iter_content(chunk_size=128):
                if not chunk: continue
                buffer += chunk.decode('utf-8', errors='ignore')
                while '\n' in buffer:
                    line, buffer = buffer.split('\n', 1)
                    if line.startswith('data: '):
                        line_data = line[len('data: '):].strip()
                        if line_data == '[DONE]': break
                        try:
                            json_data = json.loads(line_data)
                            delta = json_data['choices'][0]['delta']
                            if delta.get('reasoning_content'): reasoning_content += delta['reasoning_content']
                            if delta.get('content'): final_content += delta['content']
                        except (json.JSONDecodeError, KeyError): continue
        
        result, reason = "解析失败", "无法解析LLM返回"
        for line in final_content.strip().splitlines():
            if "用例分析结果：" in line: result = line.split("用例分析结果：")[1].strip()
            elif "用例分析内容：" in line: reason = line.split("用例分析内容：")[1].strip()
        
        debug_response = f"--- Reasoning ---\n{reasoning_content}\n\n--- Final Answer ---\n{final_content}"
        return result, reason, log_content, prompt, debug_response

    except Exception as e:
        return "API或解析错误", str(e), log_content, prompt, buffer

def handle_run(args):
    ensure_config_files_exist()
    with open(CONFIG_FILE, 'r', encoding='utf-8') as f: config = json.load(f)
    templates = load_templates()

    wb, headers = None, []
    if args.resume:
        report_path = args.resume
        if not Path(report_path).exists():
            print(f"错误: 无法找到要恢复的报告 '{report_path}'"); return
        print(f"从 {report_path} 恢复分析...")
        tasks, wb, headers = get_tasks_from_report(report_path)
    else:
        if not args.path:
            print("错误: 'run' 命令需要一个 'path' 参数。"); return
        base_path = str(Path(args.path).resolve())
        
        if args.output:
            report_path = args.output
            if not report_path.endswith(f".{args.format}"):
                report_path = f"{report_path}.{args.format}"
        else:
            report_path = f"analysis_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{args.format}"

        log_files = find_log_files(base_path, config)
        if not log_files: print(f"在 '{args.path}' 中没有找到匹配的日志文件。"); return
        
        headers, max_depth = get_dynamic_headers(log_files, base_path, args.debug)
        initialize_report(report_path, headers, log_files, max_depth, base_path)
        tasks, wb, headers = get_tasks_from_report(report_path)

    if not tasks:
        print("所有任务已完成！"); return
        
    print(f"开始使用 {args.threads} 个线程并发分析 {len(tasks)} 个文件...")
    
    file_lock = threading.Lock()
    completed_count = 0
    
    with ThreadPoolExecutor(max_workers=args.threads) as executor:
        future_to_task = {
            executor.submit(analyze_log, file_path, config, templates): (row_index, file_path)
            for row_index, file_path in tasks
        }
        
        try:
            for future in tqdm(as_completed(future_to_task), total=len(tasks), desc="Concurrent Analysis"):
                row_index, file_path = future_to_task[future]
                try:
                    result_data = future.result()
                    update_report_row(report_path, row_index, result_data, headers, file_lock, wb)
                    completed_count += 1
                    if completed_count % SAVE_INTERVAL == 0:
                        if wb:
                            with file_lock:
                                wb.save(report_path)
                except Exception as exc:
                    print(f"任务 {file_path} 生成了一个异常: {exc}")
                    error_data = ("线程异常", str(exc), "", "", "")
                    update_report_row(report_path, row_index, error_data, headers, file_lock, wb)
        finally:
            print("正在完成报告...")
            finalize_report(report_path, headers, wb)
            print(f"分析完成！报告已保存在 {report_path}")

def handle_config(args):
    ensure_config_files_exist()
    print(f"配置文件目录: {CONFIG_DIR}")
    print(f"主配置文件: {CONFIG_FILE}")
    print(f"模板文件目录: {TEMPLATES_DIR_IN_USER}")

def main():
    parser = argparse.ArgumentParser(description="AI Test Analyze: 使用LLM并发分析日志并生成报告。")
    subparsers = parser.add_subparsers(dest='command', required=True, help='Available commands')

    parser_run = subparsers.add_parser('run', help='执行分析任务')
    parser_run.add_argument("path", nargs='?', help="要分析的日志文件或目录的路径。")
    parser_run.add_argument("--output", help="输出报告的路径。如果未提供，则自动生成。")
    parser_run.add_argument("--resume", help="从指定的报告文件恢复分析。")
    parser_run.add_argument("--format", choices=['csv', 'xlsx'], default='csv', help="输出报告的格式。")
    parser_run.add_argument("--threads", type=int, default=8, help="并发分析的线程数。")
    parser_run.add_argument("--debug", action='store_true', help="在报告中包含LLM的Prompt和完整响应。")
    parser_run.set_defaults(func=handle_run)

    parser_config = subparsers.add_parser('config', help='显示配置文件路径')
    parser_config.set_defaults(func=handle_config)
    
    args = parser.parse_args()
    args.func(args)

if __name__ == "__main__":
    main()